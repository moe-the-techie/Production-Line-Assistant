import openpyxl

PRODUCT_CODES = {
    "Linear Slot Diffuser at 45Deg Angle": "LSD45^"
}
N_CORNERS = 4
POWDER_PRICE_PER_KG = 20
SHEET_SIZE = 2440 * 1220
global discount


def add_product(code, discount_inputted, quantity=1,  product_count=1, current_workbooks=None):
    """
    Adds a product to the current workbook by picking the right method for the code entered.
    :param discount_inputted: a float of the number part of the discount percentage to be applied to all products
    :param code: string representing product code
    :param quantity: integer representing the quantity of LSDs required
    :param product_count: an integer representing the number of products entered so far
    :param current_workbooks: a tuple containing two openpyxl workbooks each representing a report file to be saved
    :return:
    """

    global discount
    discount = discount_inputted

    try:
        if code == "LSD45^":

            n_slots = int(input("Enter number of slots: "))
            lsd_width = float(input("Enter width of linear slot diffuser in mm (only 16, 20, 25): "))
            lsd_length = float(input("Enter length of linear slot diffuser in mm: "))

            # Ensure all values are correct before calculating
            if n_slots <= 0 or lsd_width <= 0 or lsd_length <= 0 or lsd_width not in [16, 20, 25]:
                raise ValueError

            linear_slot_diffuser(n_slots, lsd_width, lsd_length, quantity, product_count, current_workbooks)

    except ValueError:

        if current_workbooks is None:
            print("Invalid value entered\nProcess terminated.")

        else:
            current_workbooks[0].save("Invoice.xlsx")
            current_workbooks[1].save("Internal Report.xlsx")
            print("Invalid value entered\nOutput files saved. Process terminated.")
        exit(1)


def linear_slot_diffuser(n_slots, lsd_width, lsd_length, quantity, product_count, current_workbooks):
    """
    Calculates the costs of production for the Linear Slot Diffuser and exports it into an Excel file.
    :param n_slots: integer representing the number of slots
    :param lsd_width: float representing the size of the gap in millimeters
    :param lsd_length: float representing the length of the linear slot diffuser in millimeters
    :param quantity: integer representing the quantity of LSDs required
    :param current_workbooks: a tuple containing two openpyxl workbooks each representing a report file to be saved
    :param product_count: an integer representing the number of products entered so far
    :return:
    """

    # Prices are specific for LSD per meter and changes approximately every 3 months
    outer_frame_price = 6.6
    inner_frame_price = 4.9
    louver_price = 1.6
    pipe_price = 1.2
    space_bar_price = 1.8

    # Calculations for: outer & inner frames, space bar, powder time, aluminum straps, hanging clamps, and pipe

    n_inner_frames = n_slots - 1
    inner_frame_thickness = 1.2
    outer_frame_neck_thickness = 4.4

    end_cap_size = (lsd_width + 16) * n_slots + n_inner_frames * inner_frame_thickness + 2 * outer_frame_neck_thickness

    n_pipes = int((lsd_length - 200) / 350) + 1
    pipe_size = n_pipes * end_cap_size

    space_bar_amount = lsd_length / 350 * end_cap_size - 5
    n_space_bars = pipe_size - 5
    outer_frame_size = lsd_length * 2 + end_cap_size * 2 + 380
    inner_frame_size = n_inner_frames * lsd_length

    hanging_clamp_price = 0
    n_hanging_clamps = int(lsd_length / 400)

    if n_slots == 1:
        hanging_clamp_price = 0.322580645
    elif n_slots == 2:
        hanging_clamp_price = 0.384615385
    elif n_slots == 3:
        hanging_clamp_price = 0.454545455
    elif n_slots == 4:
        hanging_clamp_price = 0.526315789
    else:
        hanging_clamp_price = 0.6

    n_aluminum_straps = round((lsd_length / 2400) * 2)
    aluminum_strap_price = 0
    aluminum_strap_type = ""

    if lsd_width == 16:
        aluminum_strap_price = 3.846153846
        aluminum_strap_type += "strap_31"

    elif lsd_width == 20:
        aluminum_strap_price = 4.54545445
        aluminum_strap_type += "strap_36"

    elif lsd_width == 25:
        aluminum_strap_price = 5
        aluminum_strap_type += "strap_40"

    louver_size = n_slots * 2 * lsd_length

    powder_weight = 0.0001333333333 * lsd_length

    powder_time = ((90 / (6000 * 12 / lsd_length)) + (90 / (6000 * 36 / lsd_length)) + 1) * 2

    # Cost calculations

    profiles_cost = (space_bar_amount * space_bar_price + inner_frame_size * inner_frame_price + outer_frame_size *
                     outer_frame_price + louver_size * louver_price + pipe_size * pipe_price) / 1000

    accessories_cost = (n_hanging_clamps * hanging_clamp_price + aluminum_strap_price * n_aluminum_straps +
                        powder_weight * POWDER_PRICE_PER_KG)

    material_cost = profiles_cost + accessories_cost

    labor_cost = material_cost * 0.4 + powder_time

    overhead_cost = labor_cost * 0.6

    total_cost = labor_cost + material_cost + overhead_cost

    unit_price = material_cost * 4
    unit_price *= 0.3

    if current_workbooks is None:
        invoice = openpyxl.load_workbook("customer template.xlsx")
        report = openpyxl.load_workbook("material template.xlsx")

    else:
        invoice = current_workbooks[0]
        report = current_workbooks[1]

    report_sheet = report.active
    invoice_sheet = invoice.active

    # Add discount percentage to the invoice
    invoice_sheet.cell(row=31, column=16).value = str(discount) + "%"

    invoice_index = product_count + 18
    report_index = product_count + 2

    if product_count > 13:
        print("Sheet full, can't add product, saving file & terminating program.")
        report.save("Internal Report.xlsx")
        invoice.save("Invoice.xlsx")
        exit(0)

    # Entering the data into the right row, notice that the 1's represent the quantity and should be updated
    # to whatever value that ends up taking on
    invoice_sheet.cell(row=invoice_index, column=2).value = PRODUCT_CODES["Linear Slot Diffuser at 45Deg Angle"]
    invoice_sheet.cell(row=invoice_index, column=3).value = "Linear Slot Diffuser at 45Deg Angle"
    invoice_sheet.cell(row=invoice_index, column=7).value = lsd_length
    invoice_sheet.cell(row=invoice_index, column=8).value = quantity
    invoice_sheet.cell(row=invoice_index, column=12).value = unit_price
    invoice_sheet.cell(row=invoice_index, column=13).value = unit_price - (unit_price * discount / 100)
    invoice_sheet.cell(row=invoice_index, column=14).value = unit_price * quantity

    report_sheet.cell(row=report_index, column=1).value = "LSD_45^"
    report_sheet.cell(row=report_index, column=2).value = quantity
    report_sheet.cell(row=report_index, column=3).value = str(round(total_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=4).value = str(round(profiles_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=5).value = str(round(accessories_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=6).value = str(round(material_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=7).value = str(round(labor_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=8).value = str(round(overhead_cost, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=9).value = str(round(outer_frame_size, 2) * quantity) + "mm"
    report_sheet.cell(row=report_index, column=10).value = str(round(outer_frame_price * outer_frame_size / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=11).value = str(round(inner_frame_size, 2) * quantity) + "mm"
    report_sheet.cell(row=report_index, column=12).value = str(round(inner_frame_price * inner_frame_size / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=13).value = str(round(louver_size, 2) * quantity) + "mm"
    report_sheet.cell(row=report_index, column=14).value = str(round(louver_price * louver_size / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=15).value = str(round(pipe_size, 2) * quantity) + "mm"
    report_sheet.cell(row=report_index, column=16).value = str(round(pipe_price * pipe_size / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=17).value = str(round(space_bar_amount, 2) * quantity) + "mm"
    report_sheet.cell(row=report_index, column=18).value = str(round(space_bar_price * space_bar_amount / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=19).value = str(round(powder_weight, 2) * quantity) + "kg"
    report_sheet.cell(row=report_index, column=20).value = str(round(powder_weight * POWDER_PRICE_PER_KG / 1000, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=21).value = str(n_hanging_clamps * quantity) + "clamps"
    report_sheet.cell(row=report_index, column=22).value = str(round(n_hanging_clamps * hanging_clamp_price, 2) * quantity) + "SAR"
    report_sheet.cell(row=report_index, column=23).value = aluminum_strap_type
    report_sheet.cell(row=report_index, column=24).value = str(n_aluminum_straps * quantity) + "straps"
    report_sheet.cell(row=report_index, column=25).value = str(round(n_aluminum_straps * aluminum_strap_price, 2) * quantity) + "SAR"


    try:
        if (input("Product Added.\nDo you wish to add another product? (y: yes, anything else: exit): ")
                .lower() in ["y", "yes"]):
            print("PRODUCT --> PRODUCT_CODE:\n")

            for product_name in PRODUCT_CODES:
                print(product_name + " --> " + PRODUCT_CODES[product_name])

            code = input("\nEnter Product Code: ")
            quantity = int(input("Enter Quantity: "))

            if code not in PRODUCT_CODES.values():
                raise ValueError

            add_product(code, discount, quantity, product_count=product_count+1, current_workbooks=(invoice, report))

        else:

            print("Saving output file and terminating process")
            report.save("Internal Report.xlsx")
            invoice.save("Invoice.xlsx")
            exit(0)

    except ValueError:
        report.save("Internal Report.xlsx")
        invoice.save("Invoice.xlsx")
        print("Invalid value entered\nOutput file saved. Process terminated.")
        exit(1)

